# -*- coding:utf-8 -*-
from csv import reader as csv_reader, writer as csv_writer
from pathlib import Path
from re import search, sub, match

from openpyxl.cell import Cell, ReadOnlyCell
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.workbook import Workbook

from .cell_style import CellStyle


def remove_end_Nones(in_list):
    """去除列表后面所有None
    :param in_list: 要处理的list
    """
    h = []
    flag = True
    for i in in_list[::-1]:
        if flag:
            if i in (None, ''):
                continue
            else:
                flag = False
        h.append(i)
    return h[::-1]


def _get_column_letter(col_idx):
    letters = []
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx, 26)
        if remainder == 0:
            remainder = 26
            col_idx -= 1
        letters.append(chr(remainder + 64))
    return ''.join(reversed(letters))


class BaseHeader(object):
    _NUM_KEY = {}
    _KEY_NUM = {}

    def __new__(cls, header=None):
        if not cls._NUM_KEY:
            for i in range(1, 18279):
                col = _get_column_letter(i)
                cls._NUM_KEY[i] = col
                cls._KEY_NUM[col] = i
        return object.__new__(cls)

    @property
    def _str_num(self):
        return Header._KEY_NUM

    @property
    def _num_str(self):
        return Header._NUM_KEY

    def __iter__(self):
        return iter(self.key_num)


class Header(BaseHeader):
    def __init__(self, header=None):
        if isinstance(header, (list, tuple)):
            self._NUM_KEY = {c: str(i) if i not in ('', None) else c
                             for c, i in enumerate(remove_end_Nones(header), start=1)}
        elif isinstance(header, dict):
            self._NUM_KEY = {c: str(v) if v not in ('', None) else None for c, v in header.items()}
        else:
            self._NUM_KEY = {}
            self._KEY_NUM = {}
            return
        self._KEY_NUM = {c: h for h, c in self._NUM_KEY.items()} if self._NUM_KEY else {}

    @property
    def key_num(self):
        return self._KEY_NUM

    @property
    def num_key(self):
        return self._NUM_KEY

    def values(self):
        return self.num_key.values()

    def items(self):
        return self.num_key.items()

    def make_row_data(self, row, row_values, None_val=None):
        """
        :param row: 行号
        :param row_values: {列序号: 值}
        :param None_val: 空值是None还是''
        :return: RowData对象
        """
        data = {self.get_key(col): val for col, val in row_values.items()}
        return RowData(row, self, None_val, data)

    def make_insert_list(self, data, auto_new, rewrite):
        """生产写入文件list格式的行数据
        :param data: 待处理行数据
        :param auto_new: 有新表头时是否自动增加
        :param rewrite: 是否需要重写表头
        :return: (处理后的行数据, 是否重写表头)
        """
        if isinstance(data, dict):
            has_int = False
            if auto_new:
                for k in data.keys():
                    if isinstance(k, int):
                        has_int = True
                    elif isinstance(k, str) and k not in self.key_num:
                        num = len(self.num_key) + 1
                        self.key_num[k] = num
                        self.num_key[num] = k
                        rewrite = True
            else:
                for k in data.keys():
                    if isinstance(k, int):
                        has_int = True
                        break

            if has_int:
                data = {self.get_num(k): v for k, v in data.items()}
                data = [data.get(c, None) for c in range(1, max(max(self.num_key), max(data)) + 1)]
            else:
                data = [data.get(c, None) for c in self.key_num]

        return data, rewrite

    def make_num_dict(self, data):
        val = {}
        for k, v in data.items():
            num = self.get_num(k)
            if num:
                val[num] = v
        return val

    def get_key(self, num):
        """返回指定列序号对应的header key，如为None返回列序号
        :param num: 列序号
        :return: 表头值或列序号
        """
        key = self[num]
        return num if key is None else key

    def get_num(self, col, is_header=True):
        """获取某列序号
        :param col: 列号、列名、表头值
        :param is_header: 当col为str时，是header的key还是列名
        :return: 列号int
        """
        if isinstance(col, int) and col > 0:
            return col
        elif isinstance(col, str):
            return self.key_num.get(col, None) if is_header else Header._KEY_NUM.get(col.upper(), None)
        else:
            raise TypeError(f'col值只能是int或str，且必须大于0。当前值：{col}')

    def __getitem__(self, item):
        if isinstance(item, str):
            self.key_num.get(item)
        elif isinstance(item, int) and item > 0:
            return self.num_key.get(item, None)
        else:
            raise ValueError('列序号不能小于1。')

    def __len__(self):
        return len(self.num_key)

    def __repr__(self):
        return str(self.num_key)

    def __bool__(self):
        return True if self.num_key else False


class ZeroHeader(Header):
    _OBJ = None

    def __new__(cls):
        super().__new__(cls)
        if cls._OBJ is None:
            cls._OBJ = object.__new__(cls)
        return cls._OBJ

    def __init__(self):
        return

    def get_num(self, col, is_header=True):
        """获取某列序号
        :param col: 列号、列名、表头值
        :param is_header: 不起实际作用
        :return: 列号int
        """
        if isinstance(col, int) and col > 0:
            return col
        elif isinstance(col, str):
            return self.key_num.get(col.upper(), None)
        else:
            raise TypeError(f'col值只能是int或str，且必须大于0。当前值：{col}')

    def make_insert_list(self, data, recorder, rewrite):
        """生产写入文件list格式的行数据
        :param data: 待处理行数据
        :param recorder: Recorder对象
        :param rewrite: 是否需要重写表头
        :return: (处理后的行数据, 是否重写表头)
        """
        if isinstance(data, dict):
            val = self.make_num_dict(data)
            data = [val.get(c, None) for c in range(1, max(val) + 1)] if val else []
        return data, False

    def __getitem__(self, item):
        return self.num_key.get(item, None) if isinstance(item, int) else self.key_num.get(item.upper(), None)

    def __len__(self):
        return 0


class RowData(dict):
    def __init__(self, row, header, None_val, seq):
        self.header = header
        self.row = row
        self.None_val = None_val
        super().__init__(seq)

    def __getitem__(self, item):
        ite = self.header[item] if isinstance(item, int) else item
        if ite is None:
            raise RuntimeError(f'header中无{item}项。\nheader：{self.header.values()}')
        return self.get(ite, self.None_val)

    def val(self, key, is_col=True, coord=False):
        """当前行获取指定列的值
        :param key: 为int时表示列序号，为str时表示列号或header key
        :param is_col: 为str时是header key还是列号
        :param coord: 为True时返回结果带坐标
        :return: coord为False时返回指定列的值，为Ture时返回(坐标, 值)
        """
        if isinstance(key, str):
            key = BaseHeader._KEY_NUM.get(key.upper(), key) if is_col else self.header[key]
        if isinstance(key, int) and key > 0:
            val = self[key]
        else:
            raise ValueError('key只能传入str或大于0的int。')
        return ((self.row, key), val) if coord else val


def align_csv(path, encoding='utf-8', delimiter=',', quotechar='"'):
    """补全csv文件，使其每行列数一样多，用于pandas读取时避免出错
    :param path: 要处理的文件路径
    :param encoding: 文件编码
    :param delimiter: 分隔符
    :param quotechar: 引用符
    :return: None
    """
    with open(path, 'r', encoding=encoding) as f:
        reader = csv_reader(f, delimiter=delimiter, quotechar=quotechar)
        lines = list(reader)
        lines_data = {}
        max_len = 0

        # 把每行列数用字典记录，并找到最长的一行
        for k, i in enumerate(lines):
            line_len = len(i)
            if line_len > max_len:
                max_len = line_len
            lines_data[k] = line_len

        # 把所有行用空值补全到和最长一行一样
        for i in lines_data:
            lines[i].extend([None] * (max_len - lines_data[i]))

        writer = csv_writer(open(path, 'w', encoding=encoding, newline=''), delimiter=delimiter, quotechar=quotechar)
        writer.writerows(lines)


def get_usable_path(path, is_file=True, parents=True):
    """检查文件或文件夹是否有重名，并返回可以使用的路径
    :param path: 文件或文件夹路径
    :param is_file: 目标是文件还是文件夹
    :param parents: 是否创建目标路径
    :return: 可用的路径，Path对象
    """
    path = Path(path)
    parent = path.parent
    if parents:
        parent.mkdir(parents=True, exist_ok=True)
    path = parent / make_valid_name(path.name)
    name = path.stem if path.is_file() else path.name
    ext = path.suffix if path.is_file() else ''

    first_time = True

    while path.exists() and path.is_file() == is_file:
        r = search(r'(.*)_(\d+)$', name)

        if not r or (r and first_time):
            src_name, num = name, '1'
        else:
            src_name, num = r.group(1), int(r.group(2)) + 1

        name = f'{src_name}_{num}'
        path = parent / f'{name}{ext}'
        first_time = None

    return path


def make_valid_name(full_name):
    """获取有效的文件名
    :param full_name: 文件名
    :return: 可用的文件名
    """
    # ----------------去除前后空格----------------
    full_name = full_name.strip()

    # ----------------去除不允许存在的字符----------------
    if search(r'[<>/\\|:*?\n"]', full_name):
        full_name = sub(r'<', '＜', full_name)
        full_name = sub(r'>', '＞', full_name)
        full_name = sub(r'/', '／', full_name)
        full_name = sub(r'\\', '＼', full_name)
        full_name = sub(r'\|', '｜', full_name)
        full_name = sub(r':', '：', full_name)
        full_name = sub(r'\*', '＊', full_name)
        full_name = sub(r'\?', '？', full_name)
        full_name = sub(r'\n', '', full_name)
        full_name = sub(r'"(.*?)"', r'“\1”', full_name)
        full_name = sub(r'"', '“', full_name)

    # ----------------使总长度不大于255个字符（一个汉字是2个字符）----------------
    r = search(r'(.*)(\.[^.]+$)', full_name)  # 拆分文件名和后缀名
    if r:
        name, ext = r.group(1), r.group(2)
        ext_long = len(ext)
    else:
        name, ext = full_name, ''
        ext_long = 0

    while get_long(name) > 255 - ext_long:
        name = name[:-1]

    return f'{name}{ext}'.rstrip('.')


def get_long(txt):
    """返回字符串中字符个数（一个汉字是2个字符）
    :param txt: 字符串
    :return: 字符个数
    """
    txt_len = len(txt)
    return int((len(txt.encode('utf-8')) - txt_len) / 2 + txt_len)


def parse_coord(coord=None, data_col=None):
    """添加数据，每次添加一行数据，可指定坐标、列号或行号
    coord只输入数字（行号）时，列号为self.data_col值，如 3；
    输入列号，或没有行号的坐标时，表示新增一行，列号为此时指定的，如'c'、',3'、(None, 3)、'None,3'；
    输入 'newline' 时，表示新增一行，列号为self.data_col值；
    输入行列坐标时，填写到该坐标，如'a3'、'3,1'、(3,1)、[3,1]；
    输入的行号可以是负数（列号不可以），代表从下往上数，-1是倒数第一行，如'a-3'、(-3, 3)
    :param coord: 坐标、列号、行号
    :param data_col: 列号，用于只传入行号的情况
    :return: 坐标tuple：(行, 列)，或(None, 列)
    """
    return_coord = None
    if not coord:  # 新增一行，列为data_col
        return_coord = None, data_col

    elif isinstance(coord, (int, float)) and coord != 0:
        return_coord = int(coord), data_col

    elif isinstance(coord, str):
        coord = coord.replace(' ', '')

        if coord.isalpha():  # 只输入列号，要新建一行
            return_coord = None, column_index_from_string(coord)

        elif ',' in coord:  # '3,1'形式
            x, y = coord.split(',')
            if x.lower() in ('', 'new', 'none', 'newline'):
                x = None
            elif x.isdigit():
                x = int(x)
            else:
                raise ValueError('行格式不正确。')

            if y.isdigit():
                y = int(y)
            elif y.isalpha():
                y = column_index_from_string(y)
            else:
                raise TypeError('列格式不正确。')

            return_coord = x, y

        else:  # 'A3'或'3A'形式
            m = match(r'^[$]?([A-Za-z]{1,3})[$]?(-?\d+)$', coord)
            if m:
                y, x = m.groups()
                return_coord = int(x), column_index_from_string(y)

            else:
                m = match(r'^[$]?(-?\d+)[$]?([A-Za-z]{1,3})$', coord)
                if not m:
                    raise ValueError(f'{coord} 坐标格式不正确。')
                x, y = m.groups()
                return_coord = int(x), column_index_from_string(y)

    elif isinstance(coord, (tuple, list)):
        if len(coord) != 2:
            raise ValueError('coord为list或tuple时长度必须为2。')

        x = None
        if coord[0] not in (None, 'new', 'newline'):
            x = int(coord[0])

        if isinstance(coord[1], int):
            y = coord[1]
        elif isinstance(coord[1], str):
            y = column_index_from_string(coord[1])
        else:
            raise TypeError('列格式不正确。')

        return_coord = x, y

    if not return_coord or return_coord[0] == 0 or return_coord[1] == 0:
        raise ValueError(f'{return_coord} 坐标格式不正确。')
    return return_coord


def process_content_xlsx(content):
    """处理单个单元格要写入的数据
    :param content: 未处理的数据内容
    :return: 处理后的数据
    """
    if isinstance(content, (str, int, float, type(None))):
        data = content
    elif isinstance(content, (Cell, ReadOnlyCell)):
        data = content.value
    else:
        data = str(content)

    if isinstance(data, str):
        data = sub(r'[\000-\010]|[\013-\014]|[\016-\037]', '', data)

    return data


def process_content_json(content):
    """处理单个单元格要写入的数据
    :param content: 未处理的数据内容
    :return: 处理后的数据
    """
    if isinstance(content, (str, int, float, type(None))):
        return content
    elif isinstance(content, (Cell, ReadOnlyCell)):
        return content.value
    else:
        return str(content)


def process_content_str(content):
    """处理单个单元格要写入的数据，以str格式输出
    :param content: 未处理的数据内容
    :return: 处理后的数据
    """
    if isinstance(content, str):
        return content
    elif content is None:
        return ''
    elif isinstance(content, (Cell, ReadOnlyCell)):
        return str(content.value)
    else:
        return str(content)


def ok_list_xlsx(data_list):
    """处理列表中数据使其符合保存规范
    :param data_list: 数据列表
    :return: 处理后的列表
    """
    if isinstance(data_list, (dict, Header)):
        data_list = data_list.values()
    return [process_content_xlsx(i) for i in data_list]


def ok_list_str(data_list):
    """处理列表中数据使其符合保存规范，所有数据都是str格式
    :param data_list: 数据列表
    :return: 处理后的列表
    """
    if isinstance(data_list, (dict, Header)):
        data_list = data_list.values()
    return [process_content_str(i) for i in data_list]


def ok_list_db(data_list):
    """处理列表中数据使其符合保存规范
    :param data_list: 数据列表
    :return: 处理后的列表
    """
    if isinstance(data_list, (dict, Header)):
        data_list = data_list.values()
    return [process_content_json(i) for i in data_list]


def get_usable_coord_int(coord, max_row, max_col):
    """返回真正写入文件的坐标
    :param coord: 已初步格式化的坐标，如(1, 2)、(None, 3)、(-3, -2)
    :param max_row: 文件最大行
    :param max_col: 文件最大列
    :return: 真正写入文件的坐标，tuple格式
    """
    row, col = coord
    if col < 0:
        col = max_col + col + 1
        if col < 1:
            raise ValueError(f'列号不能小于1。当前：{col}')

    if row is None:
        row = max_row + 1
    elif row < 0:
        row = max_row + row + 1
        if row < 1:
            raise ValueError(f'行号不能小于1。当前：{row}')

    return row, col


def get_usable_coord(coord, max_row, ws):
    """返回真正写入文件的坐标
    :param coord: 已初步格式化的坐标，如(1, 2)、(None, 3)、(-3, -2)
    :param max_row: 文件最大行
    :param ws: Worksheet对象
    :return: 真正写入文件的坐标，tuple格式
    """
    row, col = coord
    if col < 0:
        col = ws.max_column + col + 1
        if col < 1:
            raise ValueError(f'列号不能小于1。当前：{col}')

    if row is None:
        row = max_row + 1
    elif row < 0:
        row = max_row + row + 1
        if row < 1:
            raise ValueError(f'行号不能小于1。当前：{row}')

    return row, col


def data_to_list_or_dict_simplify(recorder, data):
    """将传入的数据转换为列表或字典形式，不添加前后列数据
    :param recorder: BaseRecorder对象
    :param data: 要处理的数据
    :return: 转变成列表或字典形式的数据
    """
    if data is None:
        data = tuple()
    elif not isinstance(data, (list, tuple, dict)):
        data = (data,)
    return data


def data_to_list_or_dict(recorder, data):
    """将传入的数据转换为列表或字典形式，添加前后列数据
    :param recorder: BaseRecorder对象
    :param data: 要处理的数据
    :return: 转变成列表或字典形式的数据
    """
    if data is None:
        data = tuple()
    elif not isinstance(data, (list, tuple, dict)):
        data = (data,)
    if not (recorder._before or recorder._after):
        return data

    if isinstance(data, (list, tuple)):
        return_list = []
        for i in (recorder.before, data, recorder.after):
            if isinstance(i, dict):
                return_list.extend(list(i.values()))
            elif i is None:
                pass
            elif isinstance(i, list):
                return_list.extend(i)
            elif isinstance(i, tuple):
                return_list.extend(list(i))
            else:
                return_list.extend([str(i)])

        return return_list

    elif isinstance(data, dict):
        if not recorder.before:
            pass
        elif isinstance(recorder.before, dict):
            data = {**recorder.before, **data}
        elif isinstance(recorder.before, (list, tuple)):
            data1 = list(recorder.before)
            data1.extend(data.values())
            data = data1

        if not recorder.after:
            return data

        elif isinstance(data, dict):
            if isinstance(recorder.after, dict):
                data = {**data, **recorder.after}
            elif isinstance(recorder.after, (list, tuple)):
                data = list(data)
                data.extend(recorder.after)

        elif isinstance(data, list):
            if isinstance(recorder.after, dict):
                data.extend(recorder.after.values())
            elif isinstance(recorder.after, (list, tuple)):
                data.extend(recorder.after)

        return data


def _set_style(height, styles, ws, row):
    if height is not None:
        ws.row_dimensions[row].height = height

    if styles:
        if isinstance(styles, CellStyle):
            for c in ws[row]:
                styles.to_cell(c)
        else:
            for k, s in enumerate(styles, start=1):
                if s:
                    s.to_cell(ws.cell(row=row, column=k))


def get_csv(recorder, mode):
    recorder._file_exists = True
    return (open(recorder.path, mode, newline='', encoding=recorder.encoding),
            not (recorder._file_exists or Path(recorder.path).exists()))


def get_wb(recorder):
    if recorder._file_exists or Path(recorder.path).exists():
        wb = load_workbook(recorder.path)
        new_file = False
    else:
        wb = Workbook()
        new_file = True
    recorder._file_exists = True
    return wb, new_file


def get_ws(wb, table, tables, new_file):
    new_sheet = new_file
    if table is None:
        ws = wb.active
        if ws.max_row == 1 and ws.max_column == 1 and not ws.cell(row=1, column=1).value:
            new_sheet = True

    elif table in tables:
        ws = wb[table]
        if ws.max_row == 1 and ws.max_column == 1 and not ws.cell(row=1, column=1).value:
            new_sheet = True

    elif new_file is True:
        ws = wb.active
        tables.remove(ws.title)
        ws.title = table
        tables.append(table)
        new_sheet = True

    else:
        ws = wb.create_sheet(title=table)
        tables.append(table)
        new_sheet = True

    return ws, new_sheet


def fix_openpyxl_bug(recorder, wb, ws, table):
    """尝试解决openpyxl的bug"""
    cell = ws.cell(1, 1)
    if cell.value is None:
        cell.value = ''
    wb.save(recorder.path)
    wb.close()
    wb = load_workbook(recorder.path)
    return wb, wb[table] if table else wb.active


def get_tables(path):
    wb = load_workbook(path)
    tables = wb.sheetnames
    wb.close()
    return tables


def do_nothing(*args, **kwargs):
    return


def get_key_cols(cols, header, is_header):
    """获取作为关键字的列，可以是多列
    :param cols: 列号或列名，或它们组成的list或tuple
    :param header: Header格式
    :param is_header: cols中的str表示header还是列名
    :return: 列序号列表
    """
    if cols is True:
        return True
    elif isinstance(cols, (int, str)):
        cols = header.get_num(cols, is_header)
        return [cols] if cols else []
    elif isinstance(cols, (list, tuple)):
        res = []
        for i in cols:
            i = header.get_num(i, is_header)
            if i:
                res.append(i)
        return res
    else:
        raise TypeError('col值只能是int或str。')
