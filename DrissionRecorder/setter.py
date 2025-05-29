# -*- coding:utf-8 -*-
from pathlib import Path

from openpyxl.reader.excel import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.workbook import Workbook

from .cell_style import CellStyle
from .tools import (make_valid_name, data_to_list_or_dict_simplify, data_to_list_or_dict,
                    Header, ZeroHeader, process_content_xlsx, ok_list_str)


class OriginalSetter(object):
    def __init__(self, recorder):
        self._recorder = recorder

    def cache_size(self, size):
        """设置缓存大小
        :param size: 缓存大小
        :return: 设置对象自己
        """
        if not isinstance(size, int) or size < 0:
            raise TypeError('cache_size值只能是int，且必须>=0')
        self._recorder._cache = size
        return self

    def path(self, path):
        """设置文件路径
        :param path: 文件路径
        :return: 设置对象自己
        """
        if self._recorder._path:
            self._recorder.record()
        p = Path(path)
        self._recorder._path = str(p.parent / make_valid_name(p.name))
        self._recorder._file_exists = False
        return self

    def show_msg(self, on_off):
        """设置是否显示运行信息
        :param on_off: bool表示开关
        :return: 设置对象自己
        """
        self._recorder.show_msg = on_off
        return self

    def auto_backup(self, interval=None, path=None, new_name=None):
        """设置自动备份相关参数
        :param interval: 自动保存多少次时触发备份，为0表示不自动备份，为None时不修改已设置值（初始为0）
        :param path: 备份文件存放路径，为None时不修改已设置值（初始为 'backup'）
        :param new_name: 是否用新名称备份，为None时不修改已设置值（初始为True）
        :return: 设置对象自己
        """
        if path is not None:
            self._recorder._backup_path = path
        if isinstance(new_name, bool):
            self._recorder._backup_new_name = new_name
        if interval is not None:
            self._recorder._backup_interval = interval
        return self


class BaseSetter(OriginalSetter):
    def table(self, name):
        """设置默认表名
        :param name: 表名
        :return: 设置对象自己
        """
        self._recorder._table = name
        return self

    def before(self, data):
        """设置在数据前面补充的列
        :param data: 列表、元组或字符串，为字符串时则补充一列
        :return: 设置对象自己
        """
        return self._set_after_before(True, data)

    def after(self, data):
        """设置在数据后面补充的列
        :param data: 列表、元组或字符串，为字符串时则补充一列
        :return: 设置对象自己
        """
        return self._set_after_before(False, data)

    def _set_after_before(self, before, data):
        if data is None:
            data = None
        elif isinstance(data, (list, dict)):
            data = data
        elif isinstance(data, tuple):
            data = list(data)
        else:
            data = [data]
        setattr(self._recorder, '_before' if before else '_after', data)
        if not (self._recorder._after or self._recorder._before):
            self._recorder._handle_data_method = data_to_list_or_dict_simplify
        else:
            self._recorder._handle_data_method = data_to_list_or_dict
        return self

    def encoding(self, encoding):
        """设置编码
        :param encoding: 编码格式
        :return: 设置对象自己
        """
        self._recorder._encoding = encoding
        return self


class RecorderSetter(BaseSetter):
    def header(self, header, table=None, to_file=True, row=None):
        """设置表头。只有 csv 和 xlsx 格式支持设置表头
        :param header: 表头，列表或元组
        :param table: 表名，只xlsx格式文件有效
        :param to_file: 是否写入到文件
        :param row: 指定写入文件的行号，不改变对象已设置的header_row属性，to_file为False时无效
        :return: 设置对象自己
        """
        if not header or not isinstance(header, (list, tuple)):
            raise ValueError('header不能为空且必须为list或tuple格式。')

        self._recorder.record()
        row = row or self._recorder._header_row or 1
        with self._recorder._lock:
            header = Header(header)
            if self._recorder.type == 'xlsx':
                table = table or self._recorder.table
                self._recorder._header[table] = header
                if to_file:
                    set_xlsx_header(self._recorder, header, table, row)
            elif self._recorder.type == 'csv':
                self._recorder._header[None] = header
                if to_file:
                    set_csv_header(self._recorder, header, row)

        return self

    def header_row(self, num):
        """设置标题行号
        :param num: 行号
        :return: 设置对象自己
        """
        if num < 0:
            raise ValueError('num不能小于0。')
        self._recorder._header_row = num
        self._recorder._header[self._recorder.table] = ZeroHeader() if num == 0 else None
        return self

    def delimiter(self, delimiter):
        """设置csv文件分隔符
        :param delimiter: 分隔符
        :return: 设置对象自己
        """
        self._recorder._delimiter = delimiter
        return self

    def quote_char(self, quote_char):
        """设置csv文件引用符
        :param quote_char: 引用符
        :return: 设置对象自己
        """
        self._recorder._quote_char = quote_char
        return self

    def path(self, path, file_type=None):
        """设置文件路径
        :param path: 文件路径
        :param file_type: 要设置的文件类型，为空则从文件名中获取
        :return: 设置对象自己
        """
        super().path(path)
        if not file_type:
            suffix = Path(path).suffix.lower()
            if suffix:
                file_type = suffix[1:]
        self.file_type(file_type)
        return self

    def file_type(self, file_type):
        """指定文件类型，无视文件后缀名"""
        if file_type not in self._recorder._SUPPORTS:
            file_type = 'txt'
        self._recorder._type = file_type
        self._recorder._set_methods(file_type)

        if file_type == 'xlsx' and isinstance(self._recorder._data, list):
            self._recorder._data = {self._recorder._table: self._recorder._data} if self._recorder._data else {}
        elif file_type != 'xlsx':
            if isinstance(self._recorder._data, dict):
                self._recorder._data = (self._recorder._data[self._recorder._table]
                                        if self._recorder._data.get(self._recorder._table, None) else [])
            if len(self._recorder._header) > 1:
                self._recorder._header = {None: self._recorder._header[None]}
            self._recorder._table = None

        return self

    def table(self, name):
        """设置默认表名
        :param name: 表名，为None表示使用活动表格
        :return: 设置对象自己
        """
        if isinstance(name, bool):
            name = None
        self._recorder._table = name
        return self

    def auto_new_header(self, on_off=True):
        """数据中有表头不存在的列时是否自动添加到表头，只有xlsx和csv格式有效
        :param on_off: bool表示开关
        :return: 设置对象自己
        """
        if self._recorder.type not in ('csv', 'xlsx'):
            raise TypeError('只有csv或xlsx格式可设置表头。')
        self._recorder.record()
        self._recorder._auto_new_header = on_off
        return self

    def follow_styles(self, on_off=True):
        """设置是否跟随上一行的style，只有xlsx格式有效
        :param on_off: True或False
        :return: 设置对象自己
        """
        self._recorder._follow_styles = on_off
        return self

    def default_row_height(self, height):
        """设置新行行高，只有xlsx格式有效
        :param height: 行高，传入None清空设置
        :return: 设置对象自己
        """
        self._recorder._row_height = height
        return self

    def default_styles(self, styles):
        """设置新行样式，只有xlsx格式有效，可传入多个，传入None则取消
        :param styles: 传入CellStyle对象设置整个新行，传入CellStyle对象组成的列表设置多个，传入None清空设置
        :return: 设置对象自己
        """
        self._recorder.record()
        self._recorder._styles = styles
        return self

    def data_col(self, col):
        """设置用于填充数据的列
        :param col: 列号或列名
        :return: 设置对象自己
        """
        if isinstance(col, int) and col > 0:
            self._recorder.data_col = col
        elif isinstance(col, str):
            self._recorder.data_col = column_index_from_string(col)
        else:
            raise TypeError('col值只能是int或str，且必须大于0。')
        return self

    def link_style(self, style=True):
        """设置单元格的链接样式
        :param style: CellStyle对象
        :return: 设置对象自己
        """
        if style is True:
            style = CellStyle()
            style.font.set_color("0000FF")
            style.font.set_underline('single')
        self._recorder._link_style = style
        return self


class DBSetter(BaseSetter):
    def path(self, path, table=None):
        """重写父类方法
        :param path: 文件路径
        :param table: 数据表名称
        :return: 设置对象自己
        """
        with self._recorder._lock:
            super().path(path)
            if self._recorder._conn is not None:
                self._recorder._close_connection()
            self._recorder._connect()

            if table:
                self.table(table)
            else:
                r = self._recorder.run_sql("select name from sqlite_master where type='table'")
                self._recorder._table = r[0] if r else None

            self._recorder._data = {}
            self._recorder._close_connection()
        return self

    def table(self, name):
        """设置默认表名
        :param name: 表名
        :return: 设置对象自己
        """
        if '`' in name:
            raise ValueError('table名称不能包含字符"`"。')
        self._recorder._table = name
        return self


def set_csv_header(recorder, header, row):
    """设置csv文件的表头
    :param recorder: Recorder对象
    :param header: 表头列表或元组
    :param row: 行号
    :return: None
    """
    if not recorder.path:
        raise FileNotFoundError('未指定文件。')
    from csv import writer
    if recorder._file_exists or Path(recorder.path).exists():
        with open(recorder.path, 'r', newline='', encoding=recorder._encoding) as f:
            lines = f.readlines()
            content1 = lines[:row - 1]
            content2 = lines[row:]

        with open(recorder.path, 'w', newline='', encoding=recorder._encoding) as f:
            f.write("".join(content1))
            csv_write = writer(f, delimiter=recorder._delimiter, quotechar=recorder._quote_char)
            con_len = len(content1)
            if con_len < row - 1:
                for _ in range(row - con_len - 1):
                    csv_write.writerow([])
            csv_write.writerow(ok_list_str(header))

        with open(recorder.path, 'a+', newline='', encoding=recorder._encoding) as f:
            f.write("".join(content2))

    else:
        Path(recorder.path).parent.mkdir(parents=True, exist_ok=True)
        with open(recorder.path, 'w', newline='', encoding=recorder._encoding) as f:
            csv_write = writer(f, delimiter=recorder._delimiter, quotechar=recorder._quote_char)
            for _ in range(row - 1):
                csv_write.writerow([])
            csv_write.writerow(ok_list_str(header))

    recorder._file_exists = True


def set_xlsx_header(recorder, header, table, row):
    """设置xlsx文件的表头
    :param recorder: Recorder对象
    :param header: 表头列表或元组
    :param table: 工作表名称
    :param row: 行号
    :return: None
    """
    if not recorder.path:
        raise FileNotFoundError('未指定文件。')
    if recorder._file_exists or Path(recorder.path).exists():
        wb = load_workbook(recorder.path)
        if table:
            ws = wb[table] if table in [i.title for i in wb.worksheets] else wb.create_sheet(title=table)
        else:
            ws = wb.active

    else:
        Path(recorder.path).parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        if table:
            ws.title = table

    for c, i in header.items():
        ws.cell(row, c, value=process_content_xlsx(i))
    len_row = len(ws[row])
    len_header = len(header)
    if len_row > len_header:
        for c in range(len_header + 1, len_row + 1):
            ws.cell(row, c, value=None)

    wb.save(recorder.path)
    wb.close()
    recorder._file_exists = True
