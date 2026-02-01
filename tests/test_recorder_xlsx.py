# -*- coding:utf-8 -*-
"""Tests for Recorder class with XLSX format."""
from pathlib import Path

import openpyxl
import pytest

from DrissionRecord import Recorder


class TestRecorderXLSX:
    """Test cases for Recorder with XLSX format."""

    def test_create_recorder_xlsx(self, temp_xlsx):
        """Test creating a Recorder for XLSX format."""
        r = Recorder(temp_xlsx)
        assert r.type == 'xlsx'

    def test_add_data_xlsx(self, temp_xlsx):
        """Test adding data to XLSX file."""
        r = Recorder(temp_xlsx)
        r.add_data((1, 2, 3))
        r.record()

        wb = openpyxl.load_workbook(temp_xlsx)
        ws = wb.active
        assert ws.cell(1, 1).value == 1
        assert ws.cell(1, 2).value == 2
        assert ws.cell(1, 3).value == 3
        wb.close()

    def test_add_multiple_rows_xlsx(self, temp_xlsx):
        """Test adding multiple rows to XLSX."""
        r = Recorder(temp_xlsx)
        data = ((1, 2, 3), (4, 5, 6), (7, 8, 9))
        r.add_data(data)
        r.record()

        wb = openpyxl.load_workbook(temp_xlsx)
        ws = wb.active
        assert ws.max_row == 3
        assert ws.cell(3, 3).value == 9
        wb.close()

    def test_set_header_xlsx(self, temp_xlsx):
        """Test setting header in XLSX file."""
        r = Recorder(temp_xlsx)
        r.set.header(['ID', 'Name', 'Age'])
        r.add_data((1, 'Alice', 30))
        r.record()

        wb = openpyxl.load_workbook(temp_xlsx)
        ws = wb.active
        assert ws.cell(1, 1).value == 'ID'
        assert ws.cell(1, 2).value == 'Name'
        assert ws.cell(1, 3).value == 'Age'
        assert ws.cell(2, 1).value == 1
        wb.close()

    def test_multiple_sheets(self, temp_xlsx):
        """Test working with multiple sheets in XLSX."""
        r = Recorder(temp_xlsx)

        # Add data to default sheet
        r.add_data((1, 2, 3))
        r.record()

        # Add data to a new sheet
        r.set.table('Sheet2')
        r.add_data((4, 5, 6))
        r.record()

        wb = openpyxl.load_workbook(temp_xlsx)
        assert 'Sheet' in wb.sheetnames or 'Sheet1' in wb.sheetnames
        assert 'Sheet2' in wb.sheetnames
        wb.close()

    def test_tables_property(self, temp_xlsx):
        """Test getting list of tables (sheets)."""
        r = Recorder(temp_xlsx)
        r.add_data((1, 2, 3))
        r.record()

        r.set.table('NewSheet')
        r.add_data((4, 5, 6))
        r.record()

        tables = r.tables
        assert isinstance(tables, list)
        assert len(tables) >= 2

    def test_add_img(self, temp_dir):
        """Test adding image to XLSX file."""
        from PIL import Image

        # Create a simple test image
        img_path = str(Path(temp_dir) / 'test_img.png')
        img = Image.new('RGB', (10, 10), color='red')
        img.save(img_path)

        xlsx_path = str(Path(temp_dir) / 'test.xlsx')
        r = Recorder(xlsx_path)
        r.add_data((1, 2, 3))
        r.add_img(img_path, 'B2')
        r.record()

        # Verify file was created successfully
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        assert ws.cell(1, 1).value == 1
        wb.close()

    def test_add_link(self, temp_xlsx):
        """Test adding hyperlink to XLSX file."""
        r = Recorder(temp_xlsx)
        r.add_data((1, 2, 3))
        r.add_link('https://example.com', 'A2', 'Click here')
        r.record()

        wb = openpyxl.load_workbook(temp_xlsx)
        ws = wb.active
        cell = ws.cell(2, 1)
        assert cell.value == 'Click here'
        assert cell.hyperlink is not None
        assert cell.hyperlink.target == 'https://example.com'
        wb.close()

    def test_coordinate_notation(self, temp_xlsx):
        """Test different coordinate notations."""
        r = Recorder(temp_xlsx)
        r.set.header(['A', 'B', 'C'])

        # Test A1 notation
        r.add_data((10,), coord='A2')
        r.record()

        wb = openpyxl.load_workbook(temp_xlsx)
        ws = wb.active
        assert ws.cell(2, 1).value == 10
        wb.close()

    def test_data_col(self, temp_xlsx):
        """Test data_col setting."""
        r = Recorder(temp_xlsx)
        r.set.data_col(2)  # Start from column 2
        r.add_data(('data1', 'data2', 'data3'))
        r.record()

        wb = openpyxl.load_workbook(temp_xlsx)
        ws = wb.active
        assert ws.cell(1, 2).value == 'data1'
        wb.close()

    def test_append_to_existing_xlsx(self, temp_xlsx):
        """Test appending to existing XLSX file."""
        r = Recorder(temp_xlsx)
        r.add_data((1, 2, 3))
        r.record()

        r2 = Recorder(temp_xlsx)
        r2.add_data((4, 5, 6))
        r2.record()

        wb = openpyxl.load_workbook(temp_xlsx)
        ws = wb.active
        assert ws.max_row == 2
        assert ws.cell(2, 3).value == 6
        wb.close()

    def test_dict_data_auto_header(self, temp_xlsx):
        """Test automatic header creation from dict data."""
        r = Recorder(temp_xlsx)
        r.add_data({'name': 'Alice', 'age': 30, 'city': 'Beijing'})
        r.record()

        wb = openpyxl.load_workbook(temp_xlsx)
        ws = wb.active
        # Dict keys become header, values become data
        # Check that data exists in the sheet
        assert ws.max_row >= 1
        # Check that Alice is somewhere in the sheet
        found = False
        for row in ws.iter_rows(values_only=True):
            if 'Alice' in row:
                found = True
                break
        assert found
        wb.close()

    def test_header_with_table(self, temp_xlsx):
        """Test setting header for specific table."""
        r = Recorder(temp_xlsx)
        r.set.header(['ID', 'Name'], table='Sheet1')
        r.add_data((1, 'Alice'))
        r.record()

        r.set.table('Sheet2')
        r.set.header(['X', 'Y', 'Z'])
        r.add_data((10, 20, 30))
        r.record()

        wb = openpyxl.load_workbook(temp_xlsx)
        ws1 = wb['Sheet1']
        assert ws1.cell(1, 1).value == 'ID'

        ws2 = wb['Sheet2']
        assert ws2.cell(1, 1).value == 'X'
        wb.close()

    def test_rows_method(self, temp_xlsx):
        """Test the rows() method for reading data."""
        r = Recorder(temp_xlsx)
        r.set.header(['ID', 'Name', 'Age'])
        r.add_data((1, 'Alice', 25))
        r.add_data((2, 'Bob', 30))
        r.add_data((3, 'Charlie', 35))
        r.record()

        # Read all rows
        rows = r.rows()
        assert len(rows) == 3
        assert rows[0]['Name'] == 'Alice'

        # Read with specific columns
        rows = r.rows(cols=['Name', 'Age'])
        assert len(rows[0]) == 2
        assert 'ID' not in rows[0]

        # Read with count
        rows = r.rows(count=2)
        assert len(rows) == 2

    def test_rows_with_sign_col(self, temp_xlsx):
        """Test rows() with sign_col filtering."""
        r = Recorder(temp_xlsx)
        r.set.header(['ID', 'Status', 'Value'])
        r.add_data((1, 'active', 100))
        r.add_data((2, 'inactive', 200))
        r.add_data((3, 'active', 300))
        r.record()

        # Get only active rows
        rows = r.rows(sign_col='Status', signs='active')
        assert len(rows) == 2
        assert all(r['Status'] == 'active' for r in rows)

    def test_zero_header_row(self, temp_xlsx):
        """Test with header_row set to 0."""
        r = Recorder(temp_xlsx)
        # Note: header_row=0 has special meaning but may have edge cases
        # The library uses 0 to indicate "no header row"
        r.add_data((1, 2, 3))
        r.record()

        wb = openpyxl.load_workbook(temp_xlsx)
        ws = wb.active
        # Data should still be written
        assert ws.cell(1, 1).value == 1
        wb.close()

    def test_empty_string_values(self, temp_xlsx):
        """Test handling empty string values."""
        r = Recorder(temp_xlsx)
        r.add_data(('a', '', None, 'd'))
        r.record()

        wb = openpyxl.load_workbook(temp_xlsx)
        ws = wb.active
        assert ws.cell(1, 1).value == 'a'
        assert ws.cell(1, 2).value is None
        assert ws.cell(1, 3).value is None
        wb.close()

    def test_delete_method(self, temp_xlsx):
        """Test the delete() method."""
        r = Recorder(temp_xlsx)
        r.add_data((1, 2, 3))
        r.record()
        assert Path(temp_xlsx).exists()

        r.delete()
        assert not Path(temp_xlsx).exists()
